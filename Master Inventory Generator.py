#Save the price data into an excel sheet

#Imports Here
import concurrent.futures
import datetime
import holidays
#import math
import openpyxl
from polygon import RESTClient
from openpyxl import Workbook
#RESTClient stuff here
key = '##################'
client = RESTClient(key)

#Lists here
dateList = []
tickerList = []
priceList = []

#Global Variables
executor = concurrent.futures.ThreadPoolExecutor()

#Functions
def dateGen():
    dt = datetime.date
    weekend = {5,6}
    today = datetime.date.today()
    usholiday = holidays.US()
    #add good friday dates here!
    badList = (dt(year=2018,month=12,day=5),dt(year=2012,month=10,day=30),dt(year=2012,month=10,day=29),dt(year=2022,month=4,day=15),dt(year=2021,month=4,day=2),dt(year=2020,month=4,day=10),dt(year=2019,month=4,day=19),dt(year=2018,month=3,day=30),dt(year=2017,month=4,day=14),dt(year=2016,month=3,day=25),dt(year=2015,month=4,day=3),dt(year=2014,month=4,day=18),dt(year=2013,month=3,day=29),dt(year=2012,month=4,day=6))
    for x in range(1,3650):
        minusdate = datetime.timedelta(days=x)
        todaydate = today-minusdate
        if todaydate.weekday() not in weekend:
            if todaydate not in usholiday:
                dateList.append(todaydate)
    for y in badList:
        if y in dateList:
            dateList.remove(y)
    
def tickerGen():
    today = datetime.date.today()
    path = "C:\PythonPrograms\StockPicker3\stocktickers812022.xlsx"
    #path = "C:\PythonPrograms\StockPicker2\stock_list"+str(today)+".xlsx"
    wb = openpyxl.load_workbook(filename = path)
    ws = wb['Sheet1']
    for row in ws.iter_rows():
        for cell in row:
            tickerList.append(str(cell.value))
    for item in tickerList:
        if item == None:
            tickerList.remove(item)
    tickerList.sort()

def tickerCheck(ticker):
    temporaryListTickCheck = []
    for date in dateList:
        internalListTickCheck = []
        internalListTickCheck.append(date)
        try:
            internalListTickCheck.append(client.stocks_equities_daily_open_close(ticker, str(date)).close)
        except:
            internalListTickCheck.append("HALT")
            if len(internalListTickCheck) == 2:
                temporaryListTickCheck.append(internalListTickCheck)
        else:
            if len(internalListTickCheck) == 2:
                temporaryListTickCheck.append(internalListTickCheck)
    
    priceListInternal = []
    priceListInternal.append(ticker)
    if len(temporaryListTickCheck) >= 250:
        priceListInternal.append(temporaryListTickCheck)
        priceList.append(priceListInternal)
        print(round(tickerList.index(ticker)/len(tickerList),2))

def spreadSheetAdd():
    from openpyxl import Workbook
    workBook = Workbook()
    workBookActive = workBook.active
    writeCellDate = workBookActive.cell(row=1,column=1)
    writeCellDate.value = "Date"
    rowDate = 1
    columnIndex = 1
    for dateRow in dateList:
        rowDate += 1
        writeCellDateRow = workBookActive.cell(row=rowDate,column=columnIndex)
        writeCellDateRow.value = dateRow
    for priceListAdd in priceList:
        columnIndex += 1
        rowIndex = 1
        tickerLabel = priceListAdd[0]
        writeCellTickerLabel = workBookActive.cell(row=rowIndex,column=columnIndex)
        writeCellTickerLabel.value = tickerLabel
        for priceListAddInternal in priceListAdd[1]:
            rowIndex += 1
            writeCellTickerPrice = workBookActive.cell(row=rowIndex,column=columnIndex)
            writeCellTickerPrice.value = priceListAddInternal[1]
    workBook.save(filename="Price_Master_Inventory_.xlsx")

dateGen()
dateList.reverse()
tickerGen()
dateList = tuple(dateList)
tickerList = tuple(tickerList)
executor.map(tickerCheck,tickerList)
executor.shutdown(wait=True)
priceList.sort(key = lambda priceList: priceList[0])
spreadSheetAdd()
