#Update the Master Inventory Sheet


#Imports Here
import concurrent.futures
import datetime
import holidays
#import math
import openpyxl
from polygon import RESTClient
from openpyxl import load_workbook
from stopwatch import Stopwatch

stopwatch = Stopwatch()

#Lists here
dateListNew = []
dateListOld = []
tickerList = []
priceList = []

def dateGen():
    dt = datetime.date
    weekend = {5,6}
    today = datetime.date.today()
    usholiday = holidays.US()
    #add good friday dates here!
    badList = (dt(year=2018,month=12,day=5),dt(year=2012,month=10,day=30),dt(year=2012,month=10,day=29),dt(year=2022,month=4,day=15),dt(year=2021,month=4,day=2),dt(year=2020,month=4,day=10),dt(year=2019,month=4,day=19),dt(year=2018,month=3,day=30),dt(year=2017,month=4,day=14),dt(year=2016,month=3,day=25),dt(year=2015,month=4,day=3),dt(year=2014,month=4,day=18),dt(year=2013,month=3,day=29),dt(year=2012,month=4,day=6))
    for x in range(1,3650):
        minusdate = datetime.timedelta(days=x)
        todaydate = today-minusdate
        if todaydate.weekday() not in weekend:
            if todaydate not in usholiday:
                dateListNew.append(todaydate)
    for y in badList:
        if y in dateListNew:
            dateListNew.remove(y)

def pullDates():
    wb = load_workbook(filename="Price_Master_Inventory_.xlsx")
    dataWS = wb['Data']
    maxRow = dataWS.cell(row=1,column=2).value
    newRowStart = maxRow + 1
    datePriceWS = wb['Sheet1']
    dateListOld2 = []
    for dateRow in range (2,maxRow):
        dateAdd = datePriceWS.cell(row=dateRow,column=1).value
        dateListOld2.append(dateAdd)
    dateListOld2 = set(dateListOld)
    for dateNewRemove in dateListNew:
        if dateNewRemove in dateListOld2:
            dateListNew.remove(dateNewRemove)
    
    print(dateListNew)
        

stopwatch.start()
dateGen()
pullDates()
stopwatch.stop()
print(str(stopwatch))
    
