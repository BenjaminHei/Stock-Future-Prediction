#Imports Here
import concurrent.futures
import datetime
import holidays
import math
import openpyxl
from polygon import RESTClient
import random
from stopwatch import Stopwatch
#RESTClient stuff here
key = '##############'
client = RESTClient(key)

#Lists here
dateList = []
modernList = []
ancientList = []
modernRocList = []
ancientRocList = []
tickerList = []
priceList = []
compareList = []
list1 = []
deepLearnList = []
top100 = ['TSLA','AMZN','INTC','AAPL','F','AMD','SNAP','BABA','CMCSA','VALE','ABEV','BAC','ROKU','PBR','ITUB','NVDA','META','CCL','T','NLY','BHC','GOOGL','NIO','AMCR','PLUG','MSFT','SHOP','GOOG','PLTR','XOM','UBER','AMTD','PBR-A','FCX','AAL','WBD','PFE','RIG','CSCO','BBD','VZ','SWN','TEVA','KGC','ET','NU','OXY','AVTR','X','SIRI','SOFI','KMI','PG','AMC','PSTH','ABBV','OPEN','KEY','WFC','GOLD','GGB','CVX','C','RBLX','LCID','NOK','NCLH','NEM','HBAN','RUN','PARA','MRK','VICI','TCEHY','PGY','FTI','PINS','GM','MU','BP','MRO','KO','AGNC','TTD','CVE','BMY','DKNG','CLF','AUY','COTY','VTRS','JPM','BKR','CPG','AFRM','BTG','VFC','RIVN','SLB','PYPL']
#Global Variables here
executor = concurrent.futures.ThreadPoolExecutor()
processexecutor = concurrent.futures.ProcessPoolExecutor()
stopwatch = Stopwatch()

#Functions
def dateGen():
    dt = datetime.date
    weekend = {5,6}
    today = datetime.date.today()
    usholiday = holidays.US()
    #add good friday dates here!
    badList = (dt(year=2018,month=12,day=5),dt(year=2012,month=10,day=30),dt(year=2012,month=10,day=29),dt(year=2022,month=4,day=15),dt(year=2021,month=4,day=2),dt(year=2020,month=4,day=10),dt(year=2019,month=4,day=19),dt(year=2018,month=3,day=30),dt(year=2017,month=4,day=14),dt(year=2016,month=3,day=25),dt(year=2015,month=4,day=3),dt(year=2014,month=4,day=18),dt(year=2013,month=3,day=29),dt(year=2012,month=4,day=6))
    for x in range(1,3650):
        minusdate = datetime.timedelta(days=x)
        todaydate = today-minusdate
        if todaydate.weekday() not in weekend:
            if todaydate not in usholiday:
                dateList.append(todaydate)
    for y in badList:
        if y in dateList:
            dateList.remove(y)
def tickerGen():
    today = datetime.date.today()
    path = "C:\PythonPrograms\StockPicker2\stocksymbols5172022.xlsx"
    #path = "C:\PythonPrograms\StockPicker2\stock_list"+str(today)+".xlsx"
    wb = openpyxl.load_workbook(filename = path)
    ws = wb['Sheet1']
    for row in ws.iter_rows():
        for cell in row:
            tickerList.append(str(cell.value))
    for item in tickerList:
        if item == None:
            tickerList.remove(item)
    tickerList.sort()
    
#this is a concurrent.futures function! It is different than a normal function because it
#iterates through the kwargs provided without calling them a separate thing
def tickerCheck(ticker):
    temporaryListTickCheck = []
    for date in dateList:
        internalListTickCheck = []
        internalListTickCheck.append(date)
        try:
            internalListTickCheck.append(client.stocks_equities_daily_open_close(ticker, str(date)).close)
        except:
            break
        else:
            if len(internalListTickCheck) == 2:
                temporaryListTickCheck.append(internalListTickCheck)  
    priceListInternal = []
    priceListInternal.append(ticker)
    if len(temporaryListTickCheck) >= 250:
        priceListInternal.append(temporaryListTickCheck)
        priceList.append(priceListInternal)
        print(round(tickerList.index(ticker)/len(tickerList),2))

def priceListCheck():
    for priceListChecker in priceList:
        if len(priceListChecker[1]) < len(dateList):
            priceList.remove(priceListChecker)

def separateOut():
    for priceListElement in priceList:
        if priceListElement[0] in top100:
            modernListHolder = []
            modernListHolder.append(priceListElement[0])
            modernListHolder.append([])
            while len(modernListHolder[1]) < 125:
                for dayZ in range(0,125,1):
                    modernListHolderInternal = []
                    modernListHolderInternal.append(priceListElement[1][dayZ][0])
                    modernListHolderInternal.append(priceListElement[1][dayZ][1])
                    modernListHolderInternal = tuple(modernListHolderInternal)
                    modernListHolder[1].append(modernListHolderInternal)
            modernListHolder[1] = tuple(modernListHolder[1])
            modernListHolder = tuple(modernListHolder)
            modernList.append(modernListHolder)
        ancientListHolder = []
        ancientListHolder.append(priceListElement[0])
        ancientListHolder.append([])
        day = 124
        while day <= priceListElement[1].index(priceListElement[1][-1]):
            ancientListHolderInternal = []
            ancientListHolderInternal.append(priceListElement[1][day][0])
            ancientListHolderInternal.append(priceListElement[1][day][1])
            ancientListHolderInternal = tuple(ancientListHolderInternal)
            ancientListHolder[1].append(ancientListHolderInternal)
            day +=1
        ancientListHolder[1] = tuple(ancientListHolder[1])
        ancientListHolder = tuple(ancientListHolder)
        ancientList.append(ancientListHolder)
        
def modernRoc():
    for modernListElement in modernList:
        modernRocListHolder = []
        modernRocListHolder.append(modernListElement[0])
        modernRocListInternal = []
        modernRocListInternal.append(modernListElement[1][0][0])
        modernRocListInternal.append(modernListElement[1][-1][0])
        modernRocListInternal2 = []
        for priceChangeElementModern in modernListElement[1]:
            numerator = modernListElement[1][0][1]
            denominator = priceChangeElementModern[1]
            equat = numerator / denominator
            modernRocListInternal2.append(equat)
        modernRocListInternal2 = tuple(modernRocListInternal2)  
        modernRocListInternal.append(modernRocListInternal2)
        modernRocListInternal = tuple(modernRocListInternal)
        modernRocListHolder.append(modernRocListInternal)
        modernRocListHolder = tuple(modernRocListHolder)
        modernRocList.append(modernRocListHolder)

def ancientRoc():
    for ancientListElement in ancientList:
        for dayStart in range(0,len(ancientListElement[1])-125,1):
            ancientRocListHolder = []
            ancientRocListHolder.append(ancientListElement[0])
            ancientRocListInternal = []
            ancientRocListInternal.append(ancientListElement[1][dayStart][0])
            ancientRocListInternal.append(ancientListElement[1][dayStart+125][0])
            ancientRocListInternal2 = []
            for priceChangeElementAncient in range(dayStart,dayStart+125,1):
                numerator = ancientListElement[1][dayStart][1]
                denominator = ancientListElement[1][priceChangeElementAncient][1]
                equat = numerator / denominator
                ancientRocListInternal2.append(equat)
            ancientRocListInternal2 = tuple(ancientRocListInternal2)  
            ancientRocListInternal.append(ancientRocListInternal2)
            ancientRocListInternal = tuple(ancientRocListInternal)
            ancientRocListHolder.append(ancientRocListInternal)
            ancientRocListHolder = tuple(ancientRocListHolder)
            ancientRocList.append(ancientRocListHolder)
        
def compare():
    for modernRocListElement in modernRocList:
        print(modernRocListElement[0])
        compareListHolder = []
        compareListHolder.append(modernRocListElement[0])
        compareListHolder.append(modernRocListElement[1][0])
        compareListHolder.append(modernRocListElement[1][1])
        compareListInternal = []
        for ancientRocListElement in ancientRocList:
            ancientRocListElementList = []
            ancientRocListElementList.append(ancientRocListElement[0])
            ancientRocListElementList.append(ancientRocListElement[1][0])
            ancientRocListElementList.append(ancientRocListElement[1][1])
            ancientRocListInternal = []
            for item in range(0,len(ancientRocListElement[1][2]),1):
                equat = (modernRocListElement[1][2][item]) - (ancientRocListElement[1][2][item])
                ancientRocListInternal.append(equat)
            absValue = (sum(map(abs, ancientRocListInternal)))
            ancientRocListElementList.insert(0,absValue)
            ancientRocListElementList.append(ancientRocListInternal)
            compareListInternal.append(ancientRocListElementList)
            if len(compareListInternal) > 1:
                for compareListIndex in range(0,len(compareListInternal)-1):
                    if compareListInternal[compareListIndex][1] == compareListInternal[-1][1]:
                        if compareListInternal[-1][0] < compareListInternal[compareListIndex][0]:
                            del compareListInternal[compareListIndex]
                        else:
                            del compareListInternal[-1]
            while len(compareListInternal) > 7:
                compareListInternalMaxValueList = []
                for compareListInternalMaxValue in compareListInternal:
                    compareListInternalMaxValueList.append(compareListInternalMaxValue[0])
                maxValueInternal = max(compareListInternalMaxValueList)
                for maxValueFinderInternal in compareListInternal:
                    if maxValueFinderInternal[0] == maxValueInternal:
                        compareListInternal.remove(maxValueFinderInternal)

        compareListHolderSum = 0
        for getTheAbsSum in compareListInternal:
            compareListHolderSum += getTheAbsSum[0]
        compareListHolder.insert(0,compareListHolderSum)              
        compareListHolder.append(compareListInternal)
        compareList.append(compareListHolder)
        while len(compareList) > 10:
            compareListMaxValueList = []
            for compareListMaxValue in compareList:
                compareListMaxValueList.append(compareListMaxValue[0])
            maxValue = max(compareListMaxValueList)
            for maxValueFinder in compareList:
                if maxValueFinder[0] == maxValue:
                    compareList.remove(maxValueFinder)

def pickLow():
    minimumListPickLow = []
    for compareListItem in compareList:
        minimumListPickLow.append(compareListItem[0])
    minimumValue = min(minimumListPickLow)
    for findMinimum in compareList:
        if findMinimum[0] == minimumValue:
            list1.append(findMinimum)

def deepLearn():
    for modernRocItem in modernRocList:
        if modernRocItem[0] == list1[0][1]:
            modernRocListDeepLearn = modernRocItem[1][2]
    ancientListDeep1 = []
    ancientListDeep2 = []
    ancientListDeep3 = []
    ancientListDeep4 = []
    ancientListDeep5 = []
    ancientListDeep6 = []
    ancientListDeep7 = []
    neuralNet = []
    ancientListDeepLearnList = [ancientListDeep1,ancientListDeep2,ancientListDeep3,ancientListDeep4,ancientListDeep5,ancientListDeep6,ancientListDeep7]
    for indexItem in range(0,7,1):
        ancientListDeepLearnList[indexItem].append(list1[0][4][indexItem][1])
        ancientListDeepLearnList[indexItem].append(list1[0][4][indexItem][2])
        ancientListDeepLearnList[indexItem].append(list1[0][4][indexItem][3])
        ancientListDeepLearnList[indexItem].append([])
    for ancientListDLI in ancientListDeepLearnList:
        for skimAncientRL in ancientRocList:
            if ancientListDLI[0] == skimAncientRL[0]:
                if ancientListDLI[1] == skimAncientRL[1][0]:
                    if ancientListDLI[2] == skimAncientRL[1][1]:
                        for pullRoc in skimAncientRL[1][2]:
                            ancientListDLI[3].append(pullRoc)                                          
    for ancientWeightOne in range(1,11,1):
        for ancientWeightTwo in range(1,11,1):
            for ancientWeightThree in range(1,11,1):
                for ancientWeightFour in range(1,11,1):
                    for ancientWeightFive in range(1,11,1):
                        for ancientWeightSix in range(1,11,1):
                            for ancientWeightSeven in range(1,11,1):
                                compareList = []
                                compareList2 = []
                                internalList = []
                                internalList.append(ancientWeightOne)
                                internalList.append(ancientWeightTwo)
                                internalList.append(ancientWeightThree)
                                internalList.append(ancientWeightFour)
                                internalList.append(ancientWeightFive)
                                internalList.append(ancientWeightSix)
                                internalList.append(ancientWeightSeven)
                                compareList.append(internalList)
                                for checker in range(0,125):
                                    check1 = ancientListDeepLearnList[0][3][checker] * ancientWeightOne
                                    check2 = ancientListDeepLearnList[1][3][checker] * ancientWeightTwo
                                    check3 = ancientListDeepLearnList[2][3][checker] * ancientWeightThree
                                    check4 = ancientListDeepLearnList[3][3][checker] * ancientWeightFour
                                    check5 = ancientListDeepLearnList[4][3][checker] * ancientWeightFive
                                    check6 = ancientListDeepLearnList[5][3][checker] * ancientWeightSix
                                    check7 = ancientListDeepLearnList[6][3][checker] * ancientWeightSeven
                                    numerator = check1 + check2 + check3 + check4 + check5 + check6 + check7
                                    denominator = ancientWeightOne + ancientWeightTwo + ancientWeightThree + ancientWeightFour + ancientWeightFive + ancientWeightSix + ancientWeightSeven
                                    equat = numerator / denominator
                                    equat2 = modernRocListDeepLearn[checker] - equat
                                    compareList2.append(equat2)
                                compareList.append(compareList2)
                                bigMath = 0
                                for comparisons in compareList[1]:
                                    bigMath += abs(comparisons)
                                compareList.append(bigMath)
                                if len(neuralNet) == 0:
                                    neuralNet.append(compareList)
                                else:
                                    if compareList[2] <= neuralNet[0][2]:
                                        print("New Value Added!")
                                        print(compareList[2])
                                        print("Replaced")
                                        print(neuralNet[0][2])
                                        neuralNet.clear()
                                        neuralNet.append(compareList)
    print(neuralNet)
    #deepLearnNet
    n1 = neuralNet[0][0][0] * 10
    n2 = neuralNet[0][0][1] * 10
    n3 = neuralNet[0][0][2] * 10
    n4 = neuralNet[0][0][3] * 10
    n5 = neuralNet[0][0][4] * 10
    n6 = neuralNet[0][0][5] * 10
    n7 = neuralNet[0][0][6] * 10

    #I am narrowing in the weights around the best given values to tighten it up even more
    for deepLearner1 in range(n1 - 5,n1 + 6,1):
        for deepLearner2 in range(n2 - 5,n2 + 6,1):
            for deepLearner3 in range(n3 - 5,n3 + 6,1):
                for deepLearner4 in range(n4 - 5,n4 + 6,1):
                    for deepLearner5 in range(n5 - 5,n5 + 6,1):
                        for deepLearner6 in range(n6 - 5,n6 + 6,1):
                            for deepLearner7 in range(n7 - 5,n7 + 6,1):
                                compareList = []
                                compareList2 = []
                                internalList = []
                                internalList.append(deepLearner1)
                                internalList.append(deepLearner2)
                                internalList.append(deepLearner3)
                                internalList.append(deepLearner4)
                                internalList.append(deepLearner5)
                                internalList.append(deepLearner6)
                                internalList.append(deepLearner7)
                                compareList.append(internalList)
                                for checker in range(0,125):
                                    check1 = ancientListDeepLearnList[0][3][checker] * deepLearner1
                                    check2 = ancientListDeepLearnList[1][3][checker] * deepLearner2
                                    check3 = ancientListDeepLearnList[2][3][checker] * deepLearner3
                                    check4 = ancientListDeepLearnList[3][3][checker] * deepLearner4
                                    check5 = ancientListDeepLearnList[4][3][checker] * deepLearner5
                                    check6 = ancientListDeepLearnList[5][3][checker] * deepLearner6
                                    check7 = ancientListDeepLearnList[6][3][checker] * deepLearner7
                                    numerator = check1 + check2 + check3 + check4 + check5 + check6 + check7
                                    denominator = deepLearner1 + deepLearner2 + deepLearner3 + deepLearner4 + deepLearner5 + deepLearner6 + deepLearner7
                                    equat = numerator / denominator
                                    equat2 = modernRocListDeepLearn[checker] - equat
                                    compareList2.append(equat2)
                                compareList.append(compareList2)
                                bigMath = 0
                                for comparisons in compareList[1]:
                                    bigMath += abs(comparisons)
                                compareList.append(bigMath)
                                if len(neuralNet) == 0:
                                    neuralNet.append(compareList)
                                else:
                                    if compareList[2] <= neuralNet[0][2]:
                                        print("New Value Added!")
                                        print(compareList[2])
                                        print("Replaced")
                                        print(neuralNet[0][2])
                                        neuralNet.clear()
                                        neuralNet.append(compareList)

    print(neuralNet)
    for x in range(0,10000000):
        compareList = []
        compareList2 = []
        internalList = []
        internalList.append(round(random.uniform(0.15,10.00),2))
        internalList.append(round(random.uniform(0.15,10.00),2))
        internalList.append(round(random.uniform(0.15,10.00),2))
        internalList.append(round(random.uniform(0.15,10.00),2))
        internalList.append(round(random.uniform(0.15,10.00),2))
        internalList.append(round(random.uniform(0.15,10.00),2))
        internalList.append(round(random.uniform(0.15,10.00),2))
        compareList.append(internalList)
        for checker in range(0,125):
            check1 = ancientListDeepLearnList[0][3][checker] * compareList[0][0]
            check2 = ancientListDeepLearnList[1][3][checker] * compareList[0][1]
            check3 = ancientListDeepLearnList[2][3][checker] * compareList[0][2]
            check4 = ancientListDeepLearnList[3][3][checker] * compareList[0][3]
            check5 = ancientListDeepLearnList[4][3][checker] * compareList[0][4]
            check6 = ancientListDeepLearnList[5][3][checker] * compareList[0][5]
            check7 = ancientListDeepLearnList[6][3][checker] * compareList[0][6]
            numerator = check1 + check2 + check3 + check4 + check5 + check6 + check7
            denominator = compareList[0][0] + compareList[0][1] + compareList[0][2] + compareList[0][3] + compareList[0][4] + compareList[0][5] + compareList[0][6]
            equat = numerator / denominator
            equat2 = modernRocListDeepLearn[checker] - equat
            compareList2.append(equat2)
        compareList.append(compareList2)
        bigMath = 0
        for comparisons in compareList[1]:
            bigMath += abs(comparisons)
        compareList.append(bigMath)
        if len(neuralNet) == 0:
            neuralNet.append(compareList)
        else:
            if compareList[2] <= neuralNet[0][2]:
                print("New Value Added!")
                print(compareList[2])
                print("Replaced")
                print(neuralNet[0][2])
                neuralNet.clear()
                neuralNet.append(compareList)
    list1.append(neuralNet)
    list1.append(ancientListDeepLearnList)


#REQUIRES TO DELETE FIRST LINE IN PROVIDED WORKBOOK, SOME SORT OF BUG IN THE CODE
def plotItOut():
    
    modernStockTicker = list1[0][1]
    modernStockDateStart = list1[0][2]
    modernStockDateEnd = list1[0][3]
    modernStockCapsule = [modernStockTicker,modernStockDateStart,modernStockDateEnd]
    
    ancientOneTicker = list1[0][4][0][1]
    ancientOneDateStart = list1[0][4][0][2]
    ancientOneDateEnd = list1[0][4][0][3]
    ancientOneCapsule = [ancientOneTicker,ancientOneDateStart,ancientOneDateEnd]
    
    ancientTwoTicker = list1[0][4][1][1]
    ancientTwoDateStart = list1[0][4][1][2]
    ancientTwoDateEnd = list1[0][4][1][3]
    ancientTwoCapsule = [ancientTwoTicker,ancientTwoDateStart,ancientTwoDateEnd]
    
    ancientThreeTicker = list1[0][4][2][1]
    ancientThreeDateStart = list1[0][4][2][2]
    ancientThreeDateEnd = list1[0][4][2][3]
    ancientThreeCapsule = [ancientThreeTicker,ancientThreeDateStart,ancientThreeDateEnd]
    
    ancientFourTicker = list1[0][4][3][1]
    ancientFourDateStart = list1[0][4][3][2]
    ancientFourDateEnd = list1[0][4][3][3]
    ancientFourCapsule = [ancientFourTicker,ancientFourDateStart,ancientFourDateEnd]
    
    ancientFiveTicker = list1[0][4][4][1]
    ancientFiveDateStart = list1[0][4][4][2]
    ancientFiveDateEnd = list1[0][4][4][3]
    ancientFiveCapsule = [ancientFiveTicker,ancientFiveDateStart,ancientFiveDateEnd]
    
    ancientSixTicker = list1[0][4][5][1]
    ancientSixDateStart = list1[0][4][5][2]
    ancientSixDateEnd = list1[0][4][5][3]
    ancientSixCapsule = [ancientSixTicker,ancientSixDateStart,ancientSixDateEnd]
    
    ancientSevenTicker = list1[0][4][6][1]
    ancientSevenDateStart = list1[0][4][6][2]
    ancientSevenDateEnd = list1[0][4][6][3]
    ancientSevenCapsule = [ancientSevenTicker,ancientSevenDateStart,ancientSevenDateEnd]

    plotItOutList = [ancientOneCapsule,ancientTwoCapsule,ancientThreeCapsule,ancientFourCapsule,ancientFiveCapsule,ancientSixCapsule,ancientSevenCapsule]
    for stockHistory in plotItOutList:
        tickerNew = stockHistory[0]
        dateListNew = list(dateList)
        priceListNew = []
        rocList = []
        startDate = stockHistory[1]
        finalStartDate = dateListNew[dateListNew.index(startDate)-125]
        endDate = stockHistory[2]
        goodList = []
        for date in dateListNew:
            if dateListNew.index(date) >= dateListNew.index(finalStartDate):
                if dateListNew.index(date) <= dateListNew.index(endDate):
                    goodList.append(date)
        goodList.reverse
        for dateNew in goodList:
            priceListNew.append(client.stocks_equities_daily_open_close(tickerNew, str(dateNew)).close)
        startPrice = client.stocks_equities_daily_open_close(tickerNew, str(startDate)).close
        stockHistory.append(startPrice)
        stockHistory.append(priceListNew)
        for priceToRoc in priceListNew:
            equat = startPrice / priceToRoc
            rocList.append(equat)
        rocList.reverse()
        stockHistory.append(rocList)
        
    neuralNetwork1 = list1[1][0][0][0]
    neuralNetwork2 = list1[1][0][0][1]
    neuralNetwork3 = list1[1][0][0][2]
    neuralNetwork4 = list1[1][0][0][3]
    neuralNetwork5 = list1[1][0][0][4]
    neuralNetwork6 = list1[1][0][0][5]
    neuralNetwork7 = list1[1][0][0][6]

    priceListModern = []
    rocListModern = []
    dateListModern = list(dateList)
    goodListMod = []
    for dateMod in dateListModern:
        if dateListModern.index(dateMod) >= dateListModern.index(modernStockDateStart):
                if dateListModern.index(dateMod) <= dateListModern.index(modernStockDateEnd):
                    goodListMod.append(dateMod)
    goodListMod.reverse()
    print(goodListMod)
    for dateModern in goodListMod:
        priceListModern.append(client.stocks_equities_daily_open_close(modernStockTicker, str(dateModern)).close)
    modernStockCapsule.append(priceListModern)

    deepLearnCapsule = []
    for deepIndex in range(0,250):
        deepLearn1 = plotItOutList[0][5][deepIndex] * neuralNetwork1
        deepLearn2 = plotItOutList[1][5][deepIndex] * neuralNetwork2
        deepLearn3 = plotItOutList[2][5][deepIndex] * neuralNetwork3
        deepLearn4 = plotItOutList[3][5][deepIndex] * neuralNetwork4
        deepLearn5 = plotItOutList[4][5][deepIndex] * neuralNetwork5
        deepLearn6 = plotItOutList[5][5][deepIndex] * neuralNetwork6
        deepLearn7 = plotItOutList[6][5][deepIndex] * neuralNetwork7

        numerator = deepLearn1 + deepLearn2 + deepLearn3 + deepLearn4 + deepLearn5 + deepLearn6 + deepLearn7
        denominator = neuralNetwork1 + neuralNetwork2 + neuralNetwork3 + neuralNetwork4 + neuralNetwork5 + neuralNetwork6 + neuralNetwork7
        equat = numerator / denominator
        deepLearnCapsule.append(equat)
        
    dateFinalList = []
    weekend = {5,6}
    usholiday = holidays.US()
    for dateFinal in range(0,365):
        if dateFinal == 0:
            dateFinalList.append(modernStockDateEnd)
        else:
            finalMinusDate = datetime.timedelta(days=dateFinal)
            day = modernStockDateEnd + finalMinusDate
            if day.weekday() not in weekend:
                if day not in usholiday:
                    dateFinalList.append(day)
    while len(dateFinalList) > 250:
        del dateFinalList[-1]

    from openpyxl import Workbook
    workBook = Workbook()
    workBookActive = workBook.active
    
    writeCellDate = workBookActive.cell(row=1,column=1)
    writeCellDate.value = "Date"

    writeCellModern = workBookActive.cell(row=1,column=2)
    writeCellModern.value = modernStockTicker

    writeCellDeepLearn = workBookActive.cell(row=1,column=3)
    writeCellDeepLearn.value = (modernStockTicker + " Deep Learn")
    
    mSP = modernStockCapsule[3][-1]
    
    for columns in range(4,4+len(plotItOutList)):
        writeCellAncient = workBookActive.cell(row=1,column=columns)
        writeCellAncient.value = plotItOutList[columns-4][0]
    for dateRow in range(2,2+len(dateFinalList)):
        dateZ = (dateRow - 2)
        writeCellDate2 = workBookActive.cell(row=dateRow,column=1)
        writeCellDate2.value = dateFinalList[dateZ]
    for modernRow in range(2,2+len(modernStockCapsule[3])):
        modernIndexRow = (modernRow - 2)
        writeCellModern2 = workBookActive.cell(row=modernRow,column=2)
        writeCellModern2.value = modernStockCapsule[3][modernIndexRow]
    for deepLearnRow in range(2,2+len(deepLearnCapsule)):
        deepLearnIndex = (deepLearnRow - 2)
        writeCellDeepLearn2 = workBookActive.cell(row=deepLearnRow,column=3)
        writeCellDeepLearn2.value = (mSP / deepLearnCapsule[deepLearnIndex])
    for ancientColumn in range(4,4+len(plotItOutList)):
        ancientColumnIndex = (ancientColumn - 4)
        for ancientRow in range(2,252):
            ancientRowIndex = (ancientRow - 2)
            writeCellAncient2 = workBookActive.cell(row=ancientRow,column=ancientColumn)
            writeCellAncient2.value = (mSP / plotItOutList[ancientColumnIndex][5][ancientRowIndex])
    workBook.save(filename=str(modernStockTicker)+"_Project_"+str(datetime.date.today())+".xlsx")
                     
#Commands Here
stopwatch.start()
dateGen()
tickerGen()
dateList = tuple(dateList)
tickerList = tuple(tickerList)
executor.map(tickerCheck,tickerList)
executor.shutdown(wait=True)
priceListCheck()
separateOut()
modernList = tuple(modernList)
ancientList = tuple(ancientList)
modernRoc()
modernRocList = tuple(modernRocList)
ancientRoc()
ancientRocList = tuple(ancientRocList)
compare()
compareList = tuple(compareList)
pickLow()
print(list1)
deepLearn()
print(list1)
plotItOut()
stopwatch.stop()
print(str(stopwatch))




