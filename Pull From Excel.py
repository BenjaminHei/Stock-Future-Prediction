import openpyxl
import datetime

priceList = []

def pullData():
    path = "Price_Master_Inventory_.xlsx"
    wb = openpyxl.load_workbook(filename = path)
    ws = wb['Sheet1']
    print("start")
    for column1 in range(2,60):
        internalList = []
        ticker = ws.cell(row=1,column=column1).value
        internalList.append(ticker)
        internalList2 = []
        for row1 in range(2,2492):
            internalList3 = []
            date1 = ws.cell(row=row1,column=1).value
            price1 = ws.cell(row=row1,column=column1).value
            internalList3.append(date1)
            internalList3.append(price1)
            internalList2.append(internalList3)
        internalList2.reverse()
        internalList.append(internalList2)
        priceList.append(internalList)

pullData()
print(priceList[0])
print(priceList[0][1][0][0])
print(priceList[0][1][0][1])
    

